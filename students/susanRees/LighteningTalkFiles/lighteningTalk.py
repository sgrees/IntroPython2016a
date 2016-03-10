####################
## import modules ##
####################

# the program opens the workbook by importing the openpyxl module and calling
# the openpyxl.load_workbook() function.
import openpyxl as op
wb = openpyxl.load_workbook('surveyData.xlsx')

def get_age_groups(my_sheet, my_column):
    # calls the get_sheet_by_name function to get the worksheet object
    sheet = wb.get_sheet_by_name(my_sheet)
    # Use the cell() sheet method with row and column keyword arguments
    # to pull values from all cells in one column.  puts them into a tuple.
    rawAgeData = sheet.columns[my_column]
    for cellObj in sheet.columns[my_column]:
        print(cellObj.value)
    # puts them in order
    ageDataOrdered = [sorted(rawAgeData)]
    # Using a for loop?  counts the total number of entries, creates dictionary
    # with unique entries (age) as keys, and the number of times each entry
    # appears as value?
    ageGroups = {}
    for items in ageDataOrdered:
        if items == items:
            ageGroups = {counts(items)}
            # remove duplicate items?? use itertools
        else:
            ageGroups = {counts(items)}
    # in new dictionary? iterates through dict values and divides by total
    # number of entries to provide the percentage (new value) of the whole that
    # each unique entry group (keys) represents
    # sorts entries (keys) by percentage (value) to see the order of importance
    for keys, values in (ageGroups.items()):
        numAges = len(keys)
        percentAge = int(values/numAges)
        finalNums = {percentAge}
        sorted(finalNums, percentAge)

# starts program
if __name__ == '__main__':
    my_script.main()

my_sheet = 'path to excel sheet'

# Displays them in a pie (lol) chart or bar graph of some sort?
# example from automate the borimg stuff below.
# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.get_active_sheet()
# for i in range(1, 11):         # create some data in column A
#     sheet['A' + str(i)] = i

# refObj = openpyxl.charts.Reference(sheet, (1, 1), (10, 1))

# seriesObj = openpyxl.charts.Series(refObj, title='First series')

# chartObj = openpyxl.charts.PieChart()   # can also use BarChart
# chartObj.append(seriesObj)
# chartObj.drawing.top = 50       # set the position
# chartObj.drawing.left = 100
# chartObj.drawing.width = 300    # set the size
# chartObj.drawing.height = 200

# sheet.add_chart(chartObj)
# wb.save('sampleChart.xlsx')

# convert this program to require user input (which column do you want?) or to
# run on all columns automatically?
